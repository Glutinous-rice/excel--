import openpyxl
wb=openpyxl.Workbook()   #此处的（）一定要加上，不然会在存储的时候会显示缺少参数，而且很难查出来
ws=wb['Sheet']
ws['A1']=200
ws['A2']=300
ws['A3']= '=SUM(A1:A2)'
wb.save('Sum.xlsx')