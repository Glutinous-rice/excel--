import openpyxl
from openpyxl.styles import Alignment   #此处Alignment   的A要大写
wb=openpyxl.Workbook()
ws=wb.active
ws['A1']='You are a big boy!'
ws.merge_cells('A1:B3')
center=Alignment(horizontal='center',vertical='center')
ws['A1'].alignment=center         #.alignment一定要加！
wb.save('merge.xlsx')


#########取消合并单元格用 :   ws.unmerge_cells('A1:B4')

